#!/usr/bin/python
# -*- coding:utf-8 -*-
import threading
import time
import urllib.request

import numpy
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

User_Agents = [{
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'},
    {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36 Edge/15.15063'},
    {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko'}]


def page_parser(movie_info, movie_list: list):
    title = movie_info.find( 'a', attrs={'class': 'title'} ).string.strip()
    desc = movie_info.find( 'div', attrs={'class': 'desc'} ).string.strip()
    desc_list = desc.split( '/' )
    movie_url = movie_info.find( 'a', attrs={'class': 'title'} ).get( 'href' )
    try:
        movie_origin = desc_list[0]
    except BaseException:
        movie_origin = ' 暂无'
    try:
        movie_time = str( desc_list[-5] )
    except BaseException:
        movie_time = ' 暂无'
    try:
        movie_starring = ','.join( desc_list[-4:] )
    except BaseException:
        movie_starring = ' 暂无'
    try:
        rating = movie_info.find( 'span',
                                  {'class': 'rating_nums'} ).string.strip()
    except BaseException:
        rating = '0.0'

    movie_list.append( [title, rating, movie_origin,
                        movie_time, movie_starring, movie_url] )


def movie_spider(movie_tag):
    page_num = 0
    movie_list = list()
    try_times = 0
    while True:
        url = 'https://www.douban.com/tag/' + \
              urllib.request.quote( movie_tag ) + '/movie?start=' + str( page_num * 15 )
        # Hang up the thread to avoid requesting too frequently
        time.sleep( numpy.random.rand() * 5 )
        try:
            req = requests.get(
                url, headers=User_Agents[page_num % len( User_Agents )], timeout=50 )
            req.raise_for_status()
            req.encoding = req.apparent_encoding
            source_code = req.text
            plain_text = str( source_code )
        except (requests.HTTPError, requests.URLRequired, requests.Timeout, requests.TooManyRedirects) as error:
            print( error )
            continue

        soup = BeautifulSoup( plain_text, 'lxml' )  # lxml module is required.
        list_soup = soup.find( 'div', attrs={'class': 'mod movie-list'} )
        try_times += 1
        if list_soup is None and try_times < 200:
            continue
        elif list_soup is None or len( list_soup ) <= 1:
            break  # No information returned after 200-time requesting

        for movie_info in list_soup.findAll( 'dd' ):
            page_parser( movie_info, movie_list )
            try_times = 0  # set 0 when got valid information
        page_num += 1
        print(
            "Downloading Information From Tag: {1} Page: {0} ".format(
                page_num, movie_tag ) )
    print( 'Finish Catching Tag -> {0}'.format( movie_tag ) )
    return movie_list


def fetch_list(movie_tag: str, movie_lists: list):
    movie_list = movie_spider( movie_tag )
    movie_list = sorted( movie_list, key=lambda x: x[1], reverse=True )
    movie_lists.append( movie_list )


def run_spider(movie_tag_lists):
    movie_lists = []
    threads = [
        threading.Thread(
            target=fetch_list,
            args=(
                movie_tag,
                movie_lists) ) for movie_tag in movie_tag_lists]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()
    return movie_lists


def output_to_excel(movie_lists, movie_tag_lists):
    wb = Workbook( write_only=True )
    ws = [wb.create_sheet( title=movie_tag ) for movie_tag in movie_tag_lists]
    file_name = 'Movie-List'
    for i in range( len( movie_tag_lists ) ):
        ws[i].append( ['序号', '电影名', '评分', '电影产地', '上映时间', '电影主演', '豆瓣链接'] )
        for index, movie_list in enumerate( movie_lists[i], start=1 ):
            ws[i].append( [index,
                           movie_list[0],
                           float( movie_list[1] ),
                           movie_list[2],
                           movie_list[3],
                           movie_list[4],
                           movie_list[5]] )
            print( "\r当前进度: {:.2f}%".format(
                index * 100 / len( movie_lists[i] ) ), end="" )
    for i in range( len( movie_tag_lists ) ):
        file_name += ('-' + movie_tag_lists[i])
    print()
    file_name += '.xlsx'
    wb.save( file_name )


def main():
    movie_tag_lists = ['爱情', '喜剧', '剧情', '动画', '科幻', '动作', '经典', '悬疑']
    '''
    movie_tag_lists = ['青春', '犯罪', '惊悚', '文艺', '搞笑', '纪录片', '励志', '恐怖']
    movie_tag_lists = ['战争', '短片', '魔幻', '传记', '情色', '暴力', '家庭', '音乐']
    movie_tag_lists = ['浪漫', '女性', '史诗', '童话', '黑帮', '西部', '同志']
    movie_tag_lists = ['美国', '日本', '香港', '英国', '中国', '韩国', '法国', '台湾']
    movie_tag_lists = ['中国大陆', '德国', '印度', '欧洲']
    movie_tag_lists = ['周星驰', '宫崎骏', '王家卫', '梁朝伟', 'JohnnyDepp', '尼古拉斯·凯奇', '斯皮尔伯格', '成龙']
    movie_tag_lists = ['刘德华', '张艺谋', '张国荣']
    movie_tag_lists = ['2017', '2016', '2015', '2014', '2013']
    '''
    movie_lists = run_spider( movie_tag_lists )
    output_to_excel( movie_lists, movie_tag_lists )


if __name__ == '__main__':
    main()
    print( "----All Done----" )
