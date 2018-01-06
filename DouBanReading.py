#!/usr/bin/python
# -*- coding:utf-8 -*-

import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor

import numpy
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Some User Agents  Chrome/Edge/IE
User_Agents = [{
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'},
    {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36 Edge/15.15063'},
    {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; rv:11.0) like Gecko'}]


def book_spider(book_tag):
    page_num = 0
    book_list = list()
    try_times = 0
    while True:
        url = 'https://www.douban.com/tag/' + \
              urllib.request.quote(book_tag) + '/book?start=' + str(page_num * 15)
        # Hang up the thread to avoid requesting too frequently
        time.sleep(numpy.random.rand() * 5)
        try:
            source_code = requests.get(
                url, headers=User_Agents[page_num % len(User_Agents)], timeout=50).text
            plain_text = str(source_code)
        except (requests.HTTPError, requests.URLRequired, requests.Timeout, requests.TooManyRedirects) as error:
            print(error)
            continue

        soup = BeautifulSoup(plain_text, 'lxml')  # lxml module is required.
        list_soup = soup.find('div', attrs={'class': 'mod book-list'})
        try_times += 1
        if list_soup is None and try_times < 200:
            continue
        elif list_soup is None or len(list_soup) <= 1:
            break  # No information returned after 200-time requesting

        for book_info in list_soup.findAll('dd'):
            title = book_info.find(
                'a', attrs={
                    'class': 'title'}).string.strip()
            desc = book_info.find(
                'div', attrs={
                    'class': 'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find(
                'a', attrs={'class': 'title'}).get('href')
            try:
                author_info = '/'.join(desc_list[0:-3])
            except BaseException:
                author_info = ' 暂无'
            try:
                pub_info = '/'.join(desc_list[-3:])
            except BaseException:
                pub_info = ' 暂无'
            try:
                rating = book_info.find('span',
                                        {'class': 'rating_nums'}).string.strip()
            except BaseException:
                rating = '0.0'

            book_list.append([title, rating, author_info, pub_info])
            try_times = 0  # set 0 when got valid information
        page_num += 1
        print(
            "Downloading Information From Tag: {1} Page: {0} ".format(
                page_num, book_tag))
    print('Finish Catching Tag -> {0}'.format(book_tag))
    return book_list


def fetch_list(book_tag: str, book_lists: list):
    book_list = book_spider(book_tag)
    book_list = sorted(book_list, key=lambda x: x[1], reverse=True)
    book_lists.append(book_list)


def run_spider(book_tag_lists):
    book_lists = list()
    with ThreadPoolExecutor(max_workers=len(book_tag_lists)) as executor:
        for book_tag in book_tag_lists:
            executor.submit(fetch_list, book_tag, book_lists)
    return book_lists


def output_to_excel(book_lists, book_tag_lists):
    wb = Workbook(write_only=True)
    ws = [wb.create_sheet(title=book_tag) for book_tag in book_tag_lists]
    file_name = 'Book-List'
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号', '书名', '评分', '作者/译者', '出版社', '出版日期', '售价'])
        for index, book_list in enumerate(book_lists[i], start=1):
            ws[i].append([index, book_list[0], float(
                book_list[1]), book_list[2], book_list[3]])
            print("\r当前进度: {:.2f}%".format(
                index * 100 / len(book_lists[i])), end="")
    for i in range(len(book_tag_lists)):
        file_name += ('-' + book_tag_lists[i])
    file_name += '.xlsx'
    wb.save(file_name)


if __name__ == '__main__':
    # book_tag_lists = ['心理','判断与决策','经济学','金融','商业','企业史','创业']
    # book_tag_lists = ['思想','科技','科学','股票','爱情','思维']
    # book_tag_lists = ['编程','计算机','机器学习','linux','android','数据库','互联网','算法','数据结构']
    # book_tag_lists = ['数学','应用数学','数学建模']
    # book_tag_lists = ['天文','天体物理','量子物理','弦论','黑洞']
    # book_tag_lists = ['人工智能','神经网络', '人机交互', '通信', '数据挖掘', '云计算', '物联网']  #  想爬`深度学习`标签，但是发现豆瓣此页面不存在
    # book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','健康','养生']
    # book_tag_lists = ['商业', '理财', '管理', '投资']
    # book_tag_lists = ['漫画','绘本','推理','武侠','奇幻']
    # book_tag_lists = ['名著','人文','历史','传记','哲学','诗歌','散文','港台']
    # book_tag_lists = ['科幻','科普','经典','生活','心灵','文学']
    book_tag_lists = ['个人管理', '时间管理', '投资', '文化', '宗教']
    book_lists = run_spider(book_tag_lists)
    output_to_excel(book_lists, book_tag_lists)
    print("----All Done----")
